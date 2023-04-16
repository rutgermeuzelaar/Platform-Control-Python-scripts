import pandas as pd
import openpyxl

class ReadExcelFile:
    def __init__(self, filepath: str, header_start_no: int):
        self.filepath = filepath
        self.header_start_no = header_start_no
        self.workbook = openpyxl.load_workbook(self.filepath)
        self.worksheets = self.workbook.sheetnames
        self.single_worksheet = self.workbook[
            self.worksheets[int(input(f"Which of the following worksheets do you want to work with? (0-index integer)\n"
                                      f"{self.worksheets}\n"))]]
        self.list_of_column_headers = []

        # Create a list of column headers
        non_empty = 1
        while True:
            if self.single_worksheet.cell(row=header_start_no, column=non_empty).value is None:
                break
            else:
                self.list_of_column_headers.append(
                    str(self.single_worksheet.cell(row=header_start_no, column=non_empty).value))
                non_empty += 1

        self.dataframe_dict = {header: [] for header in self.list_of_column_headers}

        # Create content lists to fill dataframe columns
        for position, (key, value) in enumerate(self.dataframe_dict.items()):
            for single_cell in range(self.header_start_no + 1, self.single_worksheet.max_row + 1):
                if self.single_worksheet.cell(row=single_cell, column=position + 1).value is None:
                    cell_value = ""
                else:
                    cell_value = str(self.single_worksheet.cell(row=single_cell, column=position + 1).value)
                value.append(cell_value)

    # Export Excel
    def export(self, export_path):
        export_dataframe = pd.DataFrame(self.dataframe_dict)
        print(export_dataframe)

        export_dataframe.to_excel(export_path)
        return export_dataframe