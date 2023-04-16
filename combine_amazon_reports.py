import re
import openpyxl
import pandas as pd
import os

# Dictionary for correct month and number.
month_name_number = {'January': 1,
                     'February': 2,
                     'March': 3,
                     'April': 4,
                     'May': 5,
                     'June': 6,
                     'July': 7,
                     'August': 8,
                     'September': 9,
                     'October': 10,
                     'November': 11,
                     'December': 12}

# Declaring column lists.
country = []
date_time = []
settlement_id = []
type_transaction = []
order_id = []
sku = []
description = []
quantity = []
marketplace = []
fulfilment = []
order_city = []
order_state = []
order_postal = []
tax_collection_model = []
product_sales = []
product_sales_tax = []
postage_credits = []
shipping_credits_tax = []
gift_wrap_credits = []
giftwrap_credits_tax = []
promotional_rebates = []
promotional_rebates_tax = []
marketplace_withheld_tax = []
selling_fees = []
fba_fees = []
other_transaction_fees = []
other = []
total = []

# List of columns.
list_of_column_lists = [
    date_time,
    settlement_id,
    type_transaction,
    order_id,
    sku,
    description,
    quantity,
    marketplace,
    fulfilment,
    order_city,
    order_state,
    order_postal,
    tax_collection_model,
    product_sales,
    product_sales_tax,
    postage_credits,
    shipping_credits_tax,
    gift_wrap_credits,
    giftwrap_credits_tax,
    promotional_rebates,
    promotional_rebates_tax,
    marketplace_withheld_tax,
    selling_fees,
    fba_fees,
    other_transaction_fees,
    other,
    total
]

# Opening the Excel file containing all the translations.
current_directory = os.path.dirname(__file__)
excel_translations_workbook = openpyxl.load_workbook(os.path.join(current_directory,
                                            'combine_amazon_reports/Payments report link vertalingen.xlsx'))

# Creating the variables for the various worksheets.
excel_translations_worksheets = excel_translations_workbook.sheetnames
excel_translations_column_header = excel_translations_workbook[excel_translations_worksheets[0]]
excel_translations_payment = excel_translations_workbook[excel_translations_worksheets[1]]

# Decide amount of English headers in sheet Translated Type of Payment.
# Needs to start at 1 because there is no column 0 in Excel.
translated_type_of_payment_counter = 1
list_of_translated_type_of_payments = []

while True:
    # Row needs to start at 2 because the first row is empty
    if excel_translations_payment.cell(row=2, column=translated_type_of_payment_counter).value is None:
        break
    else:
        translated_type_of_payment = str(excel_translations_payment.
                                         cell(row=2, column=translated_type_of_payment_counter).value)
        list_of_translated_type_of_payments.append(translated_type_of_payment)
        translated_type_of_payment_counter += 1

# Create a function that returns English column names, translated column names and translated type of payment.
def translate_payment_reports(countrycode):
    starting_row_country_dict = {'PL': 3,
                                 'ES': 4,
                                 'IT': 5,
                                 'FR': 6,
                                 'DE': 7,
                                 'NL': 8}
    english_column_names = []
    translation_list = []
    starting_row = starting_row_country_dict.get(countrycode)
    translated_transaction_type_list = []

    # Get a list of English column headers and local column headers.
    for column_header in range(2, 29):
        # Get a list of English column headers.
        if excel_translations_column_header.cell(row=2, column=column_header).value is None:
            english_value = ''
        else:
            english_value = str(excel_translations_column_header.cell(row=2, column=column_header).value)

        # Get a list of translated column headers.
        if excel_translations_column_header.cell(row=starting_row, column=column_header).value is None:
            translated_value = ''
        else:
            translated_value = str(excel_translations_column_header.cell(row=starting_row, column=column_header).value)

        english_column_names.append(english_value)
        translation_list.append(translated_value)

    # Get a list of translated payment types.
    for column_index in range(len(list_of_translated_type_of_payments)):
        if excel_translations_payment.cell(row=starting_row, column=column_index + 1).value is None:
            translated_transaction_type = ''
        else:
            translated_transaction_type = str(excel_translations_payment.
                                              cell(row=starting_row, column=column_index + 1).value)
        translated_transaction_type_list.append(translated_transaction_type)

    dictionary_english_translated = dict(zip(english_column_names, translation_list))

    return translation_list, translated_transaction_type_list, dictionary_english_translated


month = 'December'
month_number = month_name_number.get(month)
year = '2022'
client = 'client'

starting_counter = 0
# Main program loop to combine all reports and append them to the correct columns.
list_of_country_names = ['DE', 'ES', 'FR', 'IT', 'PL']
for country_code_input in list_of_country_names:

    # Loop through files and make a translation list.
    wb = openpyxl.load_workbook(os.path.join(current_directory,
                                             f'combine_amazon_reports/{client}/Date_Range_Reports_{country_code_input}.xlsx'))
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    country_translation_dict = translate_payment_reports(country_code_input)
    translation_list_for_country = country_translation_dict[0]

    dict_local_and_lists = dict(zip(translation_list_for_country, list_of_column_lists))

    print('Max row is: ' + str(ws.max_row))
    # Loop through Excel sheets to append them to the correct columns in export.
    for column_excel in range(len(translate_payment_reports(country_code_input)[0])):
        for row_excel in range(9, ws.max_row + 1):
            # Get value for each cel in the Excel sheet.
            if ws.cell(row=row_excel, column=column_excel + 1).value is None:
                value = ''
            else:
                value = str(ws.cell(row=row_excel, column=column_excel + 1).value)

            # Check the column headers against the list (in order) and append the values to the correct columns.
            name_of_column_header = str(ws.cell(row=8, column=column_excel + 1).value)

            # Check for which country the date needs to be formatted and use the correct regular expression to do so.
            if country_code_input != 'DE':
                value = re.sub(r'(\d+\d?)(\s.+\s)(\d\d\d\d)',
                               r'\1' + '-' + str(month_number) + '-' + r'\3', value)
            elif country_code_input == 'DE':
                value = re.sub(r'(\d+\d?)(\.\d+\.)(\d\d\d\d)',
                               r'\1' + '-' + str(month_number) + '-' + r'\3', value)

            try:
                dict_local_and_lists.get(name_of_column_header).append(value)
            except AttributeError:
                pass

    # Make the country column as big as the other columns, need to append here because the column will be overwritten
    # otherwise.
    for multiplier_country in range(len(date_time) - starting_counter):
        country.append(country_code_input)
    starting_counter += len(date_time) - starting_counter

    # Create dictionary with translated values.
    payment_dictionary = dict(zip(country_translation_dict[1], list_of_translated_type_of_payments))

    # Loop through type of payment and translate it.
    translated_payment_list = []

    for position, typePayment in enumerate(type_transaction):
        if payment_dictionary.get(typePayment) is None:
            pass
        else:
            type_transaction[position] = payment_dictionary.get(typePayment)

    # List for each column that needs its existing commas removed, current points replaced by commas.
    remove_comma = [product_sales, product_sales_tax, postage_credits, shipping_credits_tax, gift_wrap_credits,
                    promotional_rebates, promotional_rebates_tax, marketplace_withheld_tax, selling_fees, fba_fees,
                    other_transaction_fees, other, total]

    for main_list_comma in remove_comma:
        for position_value_comma, value_comma in enumerate(main_list_comma):
            main_list_comma[position_value_comma] = value_comma.replace(' ', '').replace('.', ',')

    report_structure = {'country': country,
                        'date/time': date_time,
                        'settlement id': settlement_id,
                        'type': type_transaction,
                        'order id': order_id,
                        'sku': sku,
                        'description': description,
                        'quantity': quantity,
                        'marketplace': marketplace,
                        'fulfilment': fulfilment,
                        'order city': order_city,
                        'order state': order_state,
                        'order postal': order_postal,
                        'tax collection model': tax_collection_model,
                        'product sales': product_sales,
                        'product sales tax': product_sales_tax,
                        'postage credits': postage_credits,
                        'shipping credits tax': shipping_credits_tax,
                        'gift wrap credits': gift_wrap_credits,
                        'gift wrap credits tax': giftwrap_credits_tax,
                        'promotional rebates': promotional_rebates,
                        'promotional rebates tax': promotional_rebates_tax,
                        'marketplace withheld tax': marketplace_withheld_tax,
                        'selling fees': selling_fees,
                        'fba fees': fba_fees,
                        'other transactions fees': other_transaction_fees,
                        'other': other,
                        'total': total
                        }

    list_with_lengths = [len(list_report_structure_value) for list_report_structure_value in report_structure.values()]
    print('Are all values of the same length?')
    if min(list_with_lengths) == max(list_with_lengths):
        print(True)
        print(report_structure['country'])
        print(min(list_with_lengths))
    else:
        for lengthOfList in list_with_lengths:
            print(False)
            print(lengthOfList)
    try:
        dataframe = pd.DataFrame(report_structure)

    # Extending lists with empty values to prevent: 'ValueError: All arrays must be of the same length'
    except ValueError:
        print('ValueError for country: ' + country_code_input)

        integer_to_multiply_empty_lists = int(max(list_with_lengths) - min(list_with_lengths))

        for empty_list_in_report_structure in report_structure.values():
            if len(empty_list_in_report_structure) != max(list_with_lengths):
                empty_list_in_report_structure.extend(['PLACEHOLDER'] * integer_to_multiply_empty_lists)
                print(len(empty_list_in_report_structure))

        dataframe = pd.DataFrame(report_structure)

    dataframe = dataframe[dataframe.type != 'Transfer']
    print(dataframe)
    main_dataframe = pd.concat([dataframe])

main_dataframe.to_excel(os.path.join(current_directory, f'combine_amazon_reports/{client}/gecombineerdrapport_{client}.xlsx'),
                                     index=False)
