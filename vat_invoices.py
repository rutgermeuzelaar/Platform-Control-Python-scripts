from mailmerge import MailMerge
from read_excel import ReadExcelFile
import openpyxl
import pandas as pd
import os

# Open the correct files.
current_directory = os.path.dirname(__file__)
vat_input_workbook = openpyxl.load_workbook(os.path.join(current_directory, 'vat_invoices/btw facturen input.xlsx'))
word_template = MailMerge(os.path.join(current_directory, 'vat_invoices/btw factuur template.docx'))

print('Merge fields are: ' + str(word_template.get_merge_fields()) + '\n')

vat_input_sheets_list = vat_input_workbook.sheetnames
vat_input_sheet = vat_input_workbook[vat_input_sheets_list[0]]

# Check the headers in the Excel file
list_of_column_headers = []
header_counter = 1
while True:
    if vat_input_sheet.cell(row=1, column=header_counter).value is None:
        break
    else:
        column_header = str(vat_input_sheet.cell(row=1, column=header_counter).value)
        list_of_column_headers.append(column_header)
        header_counter += 1

print('list_of_column_headers is: ' + str(list_of_column_headers) + '\n')

# Lists for dataframe.
factuurnummer = []
naam = []
bedrag_inc = []
datum = []
btw = []
bedrag_exc = []

dataframe_dict = {'0': datum,
                  '1': naam,
                  '2': factuurnummer,
                  '3': bedrag_inc,
                  }

list_of_lists = list(dataframe_dict.values())

for column in range(len(list_of_column_headers)):
    for row in range(2, vat_input_sheet.max_row + 1):
        if vat_input_sheet.cell(row=row, column=column + 1).value is None:
            value = ""
        else:
            value = str(vat_input_sheet.cell(row=row, column=column + 1).value)

        dataframe_dict.get(str(column)).append(value)

df = pd.DataFrame(dataframe_dict)
print(df)

for row_in_excel_file in range(25):
    naam_export = str(list_of_lists[1][row_in_excel_file])
    factuurnummer_export = str(list_of_lists[2][row_in_excel_file]).replace(' ', '')
    datum_export = str(list_of_lists[0][row_in_excel_file]).replace('00:00:00', '')

    bedrag_inc = str(list_of_lists[3][row_in_excel_file]).replace(' ', '')
    bedrag_exc = str(round(float(float(bedrag_inc) / 1.21), 2))
    btw_export = str(round(float(bedrag_inc) - float(bedrag_exc), 2))

    current_export = MailMerge(os.path.join(current_directory, 'vat_invoices/btw factuur template.docx'))
    current_export.merge(Naam=naam_export,
                    Factuurnummer=factuurnummer_export,
                    Datum=datum_export,
                    bedragEx='€ ' + bedrag_exc,
                    BTW='€ ' + btw_export,
                    bedragIn='€ ' + bedrag_inc)
    current_export.write(os.path.join(current_directory, f'vat_invoices/{factuurnummer_export}.docx'))

